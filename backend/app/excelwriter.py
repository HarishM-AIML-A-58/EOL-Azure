# excelwriter.py - Excel Export Functionality with 3-API Integration
# Uses ExcelExporter from corrected_excel.py for enhanced Excel creation

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict
import os

# Import color coding functionality
try:
    from colour import apply_color_coding_to_excel
    COLOR_CODING_AVAILABLE = True
except ImportError:
    COLOR_CODING_AVAILABLE = False
    print("[WARNING] colour.py not found. Color coding will not be applied.")


class ExcelWriter:
    """Handles Excel file creation with 3-API integration support"""

    @staticmethod
    def create_comparison(parts_data, filename, original_part):
        """Create Excel comparison file using the enhanced format from corrected_excel.py
        
        Args:
            parts_data: List of part dictionaries with merged data from Octopart, Digi-Key, Mouser
            filename: Output filename
            original_part: Original part number
        """
        
        if not parts_data:
            print("[ERROR] No data to export")
            return
        
        print(f"\n[INFO] Creating Excel comparison...")
        
        # Organize attributes
        all_attrs = set()
        for part in parts_data:
            all_attrs.update(part.keys())
        
        # Categorize attributes
        basic = ['MPN', 'Manufacturer', 'Description', 'Category']
        specs = sorted([a for a in all_attrs if a.startswith('SPEC_')])
        mouser = sorted([a for a in all_attrs if a.startswith('Mouser_')])
        
        # Get original part data for filtering
        original_part_data = parts_data[0] if parts_data else {}
        
        # Build filtered attribute list
        attributes = []
        
        # Add section header: GENERAL DETAILS
        attributes.append('=== GENERAL DETAILS ===')
        
        # Basic info
        for attr in basic:
            if attr in all_attrs:
                val = original_part_data.get(attr, 'Not Available')
                if val not in ['N/A', 'Not Available', '', None, '-']:
                    attributes.append(attr)
        
        # Add section header: SPECIFICATIONS
        if specs:
            attributes.append('=== SPECIFICATIONS ===')
            for attr in specs:
                val = original_part_data.get(attr, 'Not Available')
                if val not in ['N/A', 'Not Available', '', None, '-']:
                    # Remove SPEC_ prefix
                    clean_attr = attr.replace('SPEC_', '')
                    attributes.append(clean_attr)
        
        # Add section header: PRICING & AVAILABILITY
        if mouser:
            attributes.append('=== PRICING & AVAILABILITY ===')
            # Order Mouser fields logically
            mouser_order = [
                'Mouser_PartNumber',
                'Mouser_Stock',
                'Mouser_Availability',
                'Mouser_LeadTime'
            ]
            # Add ordered fields first with clean names
            mouser_name_map = {
                'Mouser_PartNumber': 'Mouser Part Number',
                'Mouser_Stock': 'Stock',
                'Mouser_Availability': 'Availability',
                'Mouser_LeadTime': 'LeadTime'
            }
            for mfield in mouser_order:
                if mfield in mouser:
                    attributes.append(mouser_name_map.get(mfield, mfield.replace('Mouser_', '')))
            
            # Add price breaks in sorted order
            price_fields = sorted([m for m in mouser if 'Price_Qty' in m], 
                                 key=lambda x: int(x.split('Qty')[1]))
            for pfield in price_fields:
                qty = pfield.split('Qty')[1]
                attributes.append(f'Price (Qty {qty})')
            
            # Add remaining Mouser fields
            remaining = [m for m in mouser if m not in mouser_order and 'Price_Qty' not in m]
            for mfield in remaining:
                attributes.append(mfield.replace('Mouser_', ''))
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparison"
        
        # Title
        ws['A1'] = f"Component Comparison Report - {original_part}"
        ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        ws.merge_cells(f'A1:{get_column_letter(len(parts_data) + 1)}1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Subtitle
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Parts: {len(parts_data)} | Sources: Octopart + Digi-Key + Mouser"
        ws['A2'].font = Font(italic=True, size=9, color='666666')
        ws.merge_cells(f'A2:{get_column_letter(len(parts_data) + 1)}2')
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Headers (row 3)
        ws['A3'] = 'Attribute'
        ws['A3'].font = Font(bold=True, size=11, color='FFFFFF')
        ws['A3'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
        
        for idx, part in enumerate(parts_data, 1):
            col = get_column_letter(idx + 1)
            if idx == 1:
                ws[f'{col}3'] = 'Original'
            else:
                ws[f'{col}3'] = f'Alternative {idx-1}'
            ws[f'{col}3'].font = Font(bold=True, size=11, color='FFFFFF')
            ws[f'{col}3'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            ws[f'{col}3'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        row_idx = 4
        for attr in attributes:
            # Check if section header
            if attr.startswith('==='):
                ws[f'A{row_idx}'] = attr.replace('===', '').strip()
                ws[f'A{row_idx}'].font = Font(bold=True, size=11, color='333333')
                ws[f'A{row_idx}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                ws[f'A{row_idx}'].alignment = Alignment(horizontal='left', vertical='center')
                ws.merge_cells(f'A{row_idx}:{get_column_letter(len(parts_data) + 1)}{row_idx}')
                ws.row_dimensions[row_idx].height = 25
            else:
                # Map clean attribute name back to original key
                if attr in ['MPN', 'Manufacturer', 'Description', 'Category']:
                    data_key = attr
                elif attr.startswith('Price (Qty'):
                    qty = attr.split('Qty ')[1].rstrip(')')
                    data_key = f'Mouser_Price_Qty{qty}'
                elif attr == 'Mouser Part Number':
                    data_key = 'Mouser_PartNumber'
                elif attr in ['Stock', 'Availability', 'LeadTime', 'DataSheet', 'ProductURL']:
                    data_key = f'Mouser_{attr}'
                else:
                    # Spec attribute
                    data_key = f'SPEC_{attr}'
                
                # Attribute name
                ws[f'A{row_idx}'] = attr
                ws[f'A{row_idx}'].font = Font(bold=True, size=10)
                ws[f'A{row_idx}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                
                # Data for each part
                for col_idx, part in enumerate(parts_data, 1):
                    col = get_column_letter(col_idx + 1)
                    value = part.get(data_key, 'Not Available')
                    
                    # Replace all variations with "Not Available"
                    if value in ['N/A', '-', '', None, 'Not Available']:
                        value = 'Not Available'
                    
                    ws[f'{col}{row_idx}'] = str(value)
                    ws[f'{col}{row_idx}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    ws[f'{col}{row_idx}'].font = Font(size=10)
            
            row_idx += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        for col_idx in range(2, len(parts_data) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 30
        
        # Freeze panes
        ws.freeze_panes = 'B4'
        
        # Add borders to all cells
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        for row in ws.iter_rows(min_row=3, max_row=row_idx-1, min_col=1, max_col=len(parts_data)+1):
            for cell in row:
                cell.border = thin_border
        
        # Save
        wb.save(filename)
        print(f"[SUCCESS] Excel saved: {filename}")
        print(f"   Attributes: {len([a for a in attributes if not a.startswith('===')])} (filtered)")

    def save_to_excel(self, original_part: str, recommendations: List[Dict], output_file: str, apply_color_coding: bool = True):
        """Save recommendations to Excel - compatibility method for existing app.py
        
        This method adapts the old format to work with the new ExcelExporter format.
        It converts the recommendations format and calls create_comparison.
        """
        
        # Convert recommendations format to match corrected_excel.py format
        converted_parts = []
        for rec in recommendations:
            converted = {}
            
            # Map existing fields
            if 'ManufacturerPartNumber' in rec:
                converted['MPN'] = rec['ManufacturerPartNumber']
            elif 'MPN' in rec:
                converted['MPN'] = rec['MPN']
            
            if 'Manufacturer' in rec:
                converted['Manufacturer'] = rec['Manufacturer']
            
            if 'Description' in rec:
                converted['Description'] = rec['Description']
            elif 'ShortDescription' in rec:
                converted['Description'] = rec['ShortDescription']
            
            if 'Category' in rec:
                converted['Category'] = rec['Category']
            
            # Convert all other fields - specs become SPEC_* fields
            for key, value in rec.items():
                if key not in ['ManufacturerPartNumber', 'MPN', 'Manufacturer', 'Description', 
                              'ShortDescription', 'Category', '_internal_id', '_source']:
                    # Check if it's already a spec field or should be one
                    if key.startswith('SPEC_'):
                        converted[key] = value
                    elif key.startswith('Mouser_'):
                        converted[key] = value
                    else:
                        # Regular spec field - add SPEC_ prefix
                        converted[f'SPEC_{key}'] = value
            
            converted_parts.append(converted)
        
        # Create Excel using the new format
        temp_file = output_file if not apply_color_coding else output_file.replace('.xlsx', '_temp.xlsx')
        final_file = output_file
        
        self.create_comparison(converted_parts, temp_file, original_part)
        
        # Apply color coding if requested and available
        if apply_color_coding and COLOR_CODING_AVAILABLE:
            try:
                import time
                print(f"\n[INFO] Applying color coding from colour.py...")
                time.sleep(0.5)
                apply_color_coding_to_excel(temp_file, final_file)
                
                # Clean up temp file
                if temp_file != final_file and os.path.exists(temp_file):
                    max_retries = 3
                    for attempt in range(max_retries):
                        try:
                            time.sleep(0.3)
                            os.remove(temp_file)
                            break
                        except PermissionError:
                            if attempt < max_retries - 1:
                                time.sleep(0.5)
                            else:
                                print(f"[WARNING] Could not delete temp file")
                
                print(f"\n[SAVED] Color-coded recommendations saved to: {final_file}")
            except Exception as e:
                print(f"\n[WARNING] Color coding failed: {e}")
                print(f"[INFO] Falling back to plain Excel")
                if temp_file != final_file:
                    import time
                    max_retries = 3
                    for attempt in range(max_retries):
                        try:
                            time.sleep(0.3)
                            if os.path.exists(final_file):
                                os.remove(final_file)
                            os.rename(temp_file, final_file)
                            break
                        except (PermissionError, OSError):
                            if attempt < max_retries - 1:
                                time.sleep(0.5)
                            else:
                                raise
                print(f"\n[SAVED] Recommendations saved to: {final_file}")
        else:
            if not COLOR_CODING_AVAILABLE:
                print(f"\n[INFO] Color coding not available")
            print(f"\n[SAVED] Recommendations saved to: {final_file}")
        
        print(f"Total attributes/specifications: {len([a for a in converted_parts[0].keys() if a.startswith('SPEC_')]) if converted_parts else 0}")
        print(f"[INFO] Data enriched from multiple APIs")
