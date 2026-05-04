"""
Azure OpenAI-based color coding using FFF (Form, Fit, Function) validation
"""
import os
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any, Union, Optional
from openai import AzureOpenAI


def load_fff_prompt():
    """Load FFF system prompt from file"""
    try:
        prompt_path = os.path.join(os.path.dirname(__file__), 'fff_system_prompt.md')
        with open(prompt_path, 'r', encoding='utf-8') as f:
            return f.read()
    except:
        return """You are an expert Component Engineer performing FFF (Form, Fit, Function) validation for electronic component replacement.

## Validation Rules:
- MATCH: Parameter matches exactly or within acceptable tolerance
- IMPROVED: Candidate parameter is better than EOL (higher rating, wider range)
- MINOR_DIFFERENCE: Small difference that is acceptable
- CRITICAL_FAILURE: Parameter does not meet EOL requirement
- UNKNOWN: Unable to determine due to missing or unclear data"""


def normalize_priority_map(priority_map: Union[List[Dict], List[Any]]) -> List[Dict]:
    """Convert priority_map to List[Dict] regardless of input type"""
    normalized = []
    for item in priority_map:
        if hasattr(item, 'dict'):
            # Pydantic model
            normalized.append(item.dict())
        elif hasattr(item, 'parameter') and hasattr(item, 'priority'):
            # Object with attributes
            normalized.append({
                'parameter': item.parameter,
                'priority': item.priority
            })
        elif isinstance(item, dict):
            normalized.append(item)
        else:
            # Try to convert
            normalized.append(dict(item))
    return normalized


def analyze_with_azure_fff(
    eol_specs: Dict, 
    candidate_specs: Dict, 
    priority_map: List[Dict], 
    api_key: str,
    endpoint: str,
    deployment: str,
    api_version: str
) -> Optional[Dict]:
    """Use Azure OpenAI with FFF prompt to perform color coding decisions"""
    if not api_key or not endpoint or not deployment:
        return None
    
    try:
        client = AzureOpenAI(
            api_key=api_key,
            api_version=api_version,
            azure_endpoint=endpoint
        )
        
        # Load FFF prompt
        fff_prompt = load_fff_prompt()
        
        # Build priority text
        priority_text = "\n".join([
            f"- {p.get('parameter', p.get('name', 'Unknown'))}: Priority {p.get('priority', 2)} "
            f"({'Must Match' if p.get('priority', 2) == 1 else 'Can Differ' if p.get('priority', 2) == 2 else 'Cosmetic'})"
            for p in priority_map
        ]) if priority_map else "No specific priorities defined - use default FFF rules"
        
        # Build specs text - clean up parameter names
        def clean_param(k):
            return k.replace('SPEC_', '').strip()
        
        eol_text = "\n".join([
            f"- {clean_param(k)}: {v}" 
            for k, v in eol_specs.items() 
            if v not in [None, '', 'Not Available', 'N/A', '-']
        ])
        
        candidate_text = "\n".join([
            f"- {clean_param(k)}: {v}" 
            for k, v in candidate_specs.items() 
            if v not in [None, '', 'Not Available', 'N/A', '-']
        ])
        
        # Build the comparison prompt
        user_prompt = f"""{fff_prompt}

---
COMPARISON REQUEST:

## EOL Part Specifications:
{eol_text if eol_text else "No specifications available"}

## Candidate Part Specifications:
{candidate_text if candidate_text else "No specifications available"}

## User Priority Map:
{priority_text}

---
INSTRUCTIONS:
1. Compare each parameter between EOL and Candidate parts
2. Apply FFF validation rules and priority considerations
3. Return ONLY valid JSON (no markdown, no explanation outside JSON)

Return this exact JSON structure:
{{
  "comparison_matrix": [
    {{
      "parameter": "parameter_name",
      "eol_value": "value1",
      "candidate_value": "value2",
      "ai_status": "MATCH",
      "reasoning": "explanation"
    }}
  ],
  "overall_status": "MATCH"
}}

Valid ai_status values: "MATCH", "IMPROVED", "MINOR_DIFFERENCE", "CRITICAL_FAILURE", "UNKNOWN"
Valid overall_status values: "MATCH", "IMPROVED", "MINOR_DIFFERENCE", "CRITICAL_FAILURE"
"""
        
        response = client.chat.completions.create(
            model=deployment,
            messages=[
                {"role": "system", "content": "You are a professional component engineer."},
                {"role": "user", "content": user_prompt}
            ],
            response_format={"type": "json_object"}
        )
        
        # Parse JSON
        result = json.loads(response.choices[0].message.content)
        
        print(f"[OK] Azure OpenAI analysis completed - {len(result.get('comparison_matrix', []))} parameters analyzed")
        return result
    
    except Exception as e:
        print(f"[WARNING] Azure OpenAI FFF analysis error: {e}")
        return None


def apply_azure_color_coding_to_excel(
    input_file: str, 
    output_file: str, 
    eol_part_data: Dict, 
    priority_map: Union[List[Dict], List[Any]], 
    azure_creds: Dict
):
    """
    Apply Azure OpenAI-based color coding to Excel using FFF validation
    """
    
    print("=" * 80)
    print("APPLYING AZURE OPENAI FFF COLOR CODING TO EXCEL")
    print("=" * 80)
    
    # Normalize priority_map to List[Dict]
    normalized_priority = normalize_priority_map(priority_map) if priority_map else []
    
    # Load the workbook
    wb = load_workbook(input_file)
    ws = wb.active
    
    # Define color fills
    yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Define fonts and alignment
    bold_font = Font(bold=True, size=11)
    header_font = Font(bold=True, size=14)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Get dimensions
    max_row = ws.max_row
    max_col = ws.max_column
    
    # Structure
    header_row = 3
    data_start_row = 4
    
    # Read reference values
    reference_values = {}
    for row_idx in range(data_start_row, max_row + 1):
        attribute_name = ws.cell(row=row_idx, column=1).value
        if attribute_name and not str(attribute_name).startswith('==='):
            reference_values[attribute_name] = ws.cell(row=row_idx, column=2).value
    
    # Apply initial formatting
    for col_idx in range(2, max_col + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = yellow_fill
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx in range(header_row, max_row + 1):
        cell = ws.cell(row=row_idx, column=1)
        cell.fill = gray_fill
        cell.font = bold_font
        cell.alignment = left_align
        cell.border = thin_border

    # Color EOL column green
    for row_idx in range(data_start_row, max_row + 1):
        if ws.cell(row=row_idx, column=1).value:
            cell = ws.cell(row=row_idx, column=2)
            cell.fill = green_fill
            cell.alignment = left_align
            cell.border = thin_border
    
    # Statistics
    comparison_stats = {'match': 0, 'improved': 0, 'minor_difference': 0, 'critical_failure': 0, 'unknown': 0}
    
    # Process candidates
    for col_idx in range(3, max_col + 1):
        candidate_specs = {}
        for row_idx in range(data_start_row, max_row + 1):
            attr = ws.cell(row=row_idx, column=1).value
            if attr and not str(attr).startswith('==='):
                spec_key = f'SPEC_{attr}' if not str(attr).startswith('SPEC_') else attr
                candidate_specs[spec_key] = ws.cell(row=row_idx, column=col_idx).value
        
        # Get Azure analysis
        result = analyze_with_azure_fff(
            eol_part_data, candidate_specs, normalized_priority,
            api_key=azure_creds.get('api_key'),
            endpoint=azure_creds.get('endpoint'),
            deployment=azure_creds.get('deployment'),
            api_version=azure_creds.get('api_version')
        )
        
        if result:
            gemini_map = {} # keeping var name for logic reuse
            for item in result.get('comparison_matrix', []):
                param = item.get('parameter', '')
                status = item.get('ai_status', 'UNKNOWN').upper()
                gemini_map[param] = status
                gemini_map[param.replace('SPEC_', '')] = status
                gemini_map[f'SPEC_{param}'] = status
            
            for row_idx in range(data_start_row, max_row + 1):
                attr = ws.cell(row=row_idx, column=1).value
                if not attr or str(attr).startswith('==='): continue
                
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = left_align
                cell.border = thin_border
                
                status = gemini_map.get(attr) or gemini_map.get(f'SPEC_{attr}') or gemini_map.get(str(attr).replace('SPEC_', '')) or 'UNKNOWN'
                
                if status == 'MATCH':
                    cell.fill = green_fill
                    comparison_stats['match'] += 1
                elif status == 'IMPROVED':
                    cell.fill = light_green_fill
                    comparison_stats['improved'] += 1
                elif status == 'MINOR_DIFFERENCE':
                    cell.fill = orange_fill
                    comparison_stats['minor_difference'] += 1
                elif status == 'CRITICAL_FAILURE':
                    cell.fill = red_fill
                    comparison_stats['critical_failure'] += 1
                else:
                    if cell.value in [None, '', 'Not Available', 'N/A', '-']:
                        cell.fill = red_fill
                        comparison_stats['critical_failure'] += 1
                    else:
                        cell.fill = orange_fill
                        comparison_stats['unknown'] += 1
        else:
            # Fallback
            for row_idx in range(data_start_row, max_row + 1):
                attr = ws.cell(row=row_idx, column=1).value
                if not attr or str(attr).startswith('==='): continue
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = left_align
                cell.border = thin_border
                ref_val = reference_values.get(attr)
                curr_val = cell.value
                if curr_val in [None, '', 'N/A', '-']:
                    cell.fill = red_fill
                    comparison_stats['critical_failure'] += 1
                elif str(ref_val).strip().lower() == str(curr_val).strip().lower():
                    cell.fill = green_fill
                    comparison_stats['match'] += 1
                else:
                    cell.fill = orange_fill
                    comparison_stats['minor_difference'] += 1
    
    # Finalize
    ws.column_dimensions['A'].width = 35
    for col_idx in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 25
    
    wb.save(output_file)
    wb.close()
    print(f"[SUCCESS] Azure OpenAI color coding applied to {output_file}")


def apply_color_coding_to_excel(input_file: str, output_file: str):
    """
    Standard color coding fallback (non-AI).
    Maintains compatibility with legacy calls.
    """
    return apply_azure_color_coding_to_excel(
        input_file=input_file,
        output_file=output_file,
        eol_part_data={},
        priority_map=[],
        azure_creds={}
    )
